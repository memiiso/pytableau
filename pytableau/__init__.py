#! /usr/bin/env python
# -*- coding: utf-8 -*-
import csv
import logging
import os
import shutil
import smtplib
import sys
import tempfile
import time
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from os.path import basename
from pathlib import Path
from smtpd import COMMASPACE
from urllib.parse import quote_plus

import pandas as pd
import tableaudocumentapi
import tableauserverclient as TSC
from PIL import Image
from openpyxl import Workbook
from tableauserverclient import ViewItem, WorkbookItem, DatasourceItem, ProjectItem, PDFRequestOptions, \
    ImageRequestOptions, CSVRequestOptions

try:
    import PyPDF3
except ImportError:
    raise Exception('Please `pip install PyPDF3` to use this module')

log = logging.getLogger('PyTableau')
log.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
log.addHandler(handler)


class PyTableauUtils():
    """

    """

    @staticmethod
    def NoneToStr(string):
        """

        :param string:
        :return:
        """
        return (string or '').replace('\r\n', ' ').replace('\n', ' ').replace('\t', ' ')

    @staticmethod
    def clean_folder(dirPath):
        """

        :param dirPath:
        """
        fileList = os.listdir(dirPath)
        for fileName in fileList:
            file = dirPath + "/" + fileName
            if os.path.isdir(file):
                shutil.rmtree(file)
            else:
                os.remove(file)


class PyTableau():
    """

    """

    def __init__(self, server_address, username, password, site_id, use_server_version=True, verify_ssl=True):

        self.tableau_auth = TSC.TableauAuth(username=username, password=password, site_id=site_id)
        self.server = TSC.Server(server_address=server_address, use_server_version=use_server_version)

        self.server.add_http_options({'verify': verify_ssl})
        log.info("Tableau Server Address %s " % self.server.server_address)
        log.info("Tableau Server Baseurl %s " % self.server.baseurl)
        self.sign_in()

    def __del__(self):
        try:
            self.server.auth.sign_out()
        except:
            pass

    def sign_in(self):
        self.tableau_auth = self.server.auth.sign_in(self.tableau_auth)

    def sign_out(self):
        self.server.auth.sign_out()

    def download_all_datasources(self, download_dir, include_extract=False):
        """

        :param download_dir:
        :param include_extract:
        """
        PyTableauUtils.clean_folder(download_dir)
        for server_datasource in TSC.Pager(self.server.datasources):
            ds_download_dir = os.path.join(download_dir, server_datasource.project_name)
            if not os.path.exists(ds_download_dir):
                os.makedirs(ds_download_dir)
            path = self.server.datasources.download(server_datasource.id, filepath=ds_download_dir,
                                                    include_extract=include_extract)
            log.info("Downloaded datasource: %s " % path)
        log.info("Download Completed! Download directory %s" % download_dir)

    def download_all_workbooks(self, download_dir):
        """
        download all workbooks from server to given directory

        :param download_dir:
        """

        PyTableauUtils.clean_folder(download_dir)
        for server_workbook in TSC.Pager(self.server.workbooks):
            wb_download_dir = os.path.join(download_dir, server_workbook.project_name)
            if not os.path.exists(wb_download_dir):
                os.makedirs(wb_download_dir)
            try:
                path = self.server.workbooks.download(server_workbook.id, filepath=wb_download_dir,
                                                      include_extract=False)
                log.info("Downloaded workbook: %s " % path)
            except Exception:
                # pass this error "UnicodeEncodeError 'ascii' codec can't encode characters in position : ordinal
                # not in range(128)"
                log.info("Skipping workbook %s " % server_workbook.name)
        log.info("Download Completed! Download directory %s " % download_dir)

    def get_all_workbook_fields(self, workbooks_dir):
        """
        get all fields from workbooks found in given directory

        :param workbooks_dir:
        :return:
        """
        log.info("Extracting all workbook fields")
        rows_list = []

        for root, _, filenames in os.walk(workbooks_dir):
            for filename in filenames:
                if "twb" in filename:
                    # read metadata of workbook
                    try:
                        log.info("Processing %s " % filename)
                        my_wb = tableaudocumentapi.workbook.Workbook(os.path.join(root, filename))
                    except Exception:
                        continue
                    for myDS in my_wb.datasources:
                        for _, field in myDS.fields.items():
                            if len(field.worksheets) > 0:
                                for myWorksheet in field.worksheets:
                                    field_dict = self._field_dict(field, datasource=myDS, workbook=my_wb,
                                                                  worksheet=myWorksheet)
                                    rows_list.append(field_dict)
                            else:
                                field_dict = self._field_dict(field, datasource=myDS, workbook=my_wb)
                                rows_list.append(field_dict)
        return pd.DataFrame(rows_list)

    def get_all_datasource_fields(self, datasource_dir):
        """

        :param datasource_dir:
        :return:
        """
        rows_list = []
        for root, _, filenames in os.walk(datasource_dir):
            for filename in filenames:
                if "tds" in filename:
                    log.info("Processing " + filename)
                    # read metadata of workbook
                    my_ds = tableaudocumentapi.datasource.Datasource.from_file(os.path.join(root, filename))
                    for _, field in my_ds.fields.items():
                        field_dict = self._field_dict(field, datasource=my_ds)
                        rows_list.append(field_dict)

        return pd.DataFrame(rows_list)

    def export_all_workbook_fields_to_csv(self, workbooks_dir):
        """

        :param workbooks_dir:
        """
        self.download_all_workbooks(download_dir=workbooks_dir)
        field_list_file = workbooks_dir + '/all_workbook_fields.csv'
        log.info("Putting all workbook fields into %s " % field_list_file)
        df_wb_fields = self.get_all_workbook_fields(workbooks_dir)
        df_wb_fields.to_csv(field_list_file, sep='\t', encoding='utf-8')
        log.info("Created %s " % field_list_file)

    def export_all_datasource_fields_to_csv(self, datasource_dir):
        """

        :param datasource_dir:
        """
        self.download_all_workbooks(download_dir=datasource_dir)
        field_list_file = datasource_dir + '/all_datasource_fields.csv'
        log.info("Putting all datasource fields into %s " % (field_list_file,))
        df_wb_fields = self.get_all_datasource_fields(datasource_dir)
        df_wb_fields.to_csv(field_list_file, sep='\t', encoding='utf-8')
        log.info("Created %s" % field_list_file)

    def _field_dict(self, field, datasource, workbook=None, worksheet=None):
        """

        :param field:
        :param datasource:
        :param workbook:
        :param worksheet:
        :return:
        """
        row = {"field_name": PyTableauUtils.NoneToStr(field.caption),
               "field_aggregation": PyTableauUtils.NoneToStr(field._aggregation),
               "field_alias": PyTableauUtils.NoneToStr(field.alias),
               "field_calculation": PyTableauUtils.NoneToStr(field.calculation),
               "field_datatype": PyTableauUtils.NoneToStr(field.datatype),
               "field_description": PyTableauUtils.NoneToStr(field.description),
               "field_id": PyTableauUtils.NoneToStr(field.id),
               "field_role": PyTableauUtils.NoneToStr(field.role), "field_type": PyTableauUtils.NoneToStr(field._type),
               'data_source_name': PyTableauUtils.NoneToStr(datasource.name or datasource.caption),
               'data_source_caption': PyTableauUtils.NoneToStr(datasource.caption),
               'data_source_version': PyTableauUtils.NoneToStr(datasource.version), 'data_source_connections': '@TODO'}

        if workbook:
            row['workbook_name'] = PyTableauUtils.NoneToStr(os.path.basename(workbook.filename))
        else:
            row['workbook_name'] = ''

        if worksheet:
            row['worksheet_name'] = PyTableauUtils.NoneToStr(worksheet)
        else:
            row['workbook_name'] = ''

        return row

    def refresh_extracts(self, datasource_names, retry_attempt=2, synchronous=False):
        """

        :param synchronous:
        :param datasource_names:
        :param retry_attempt:
        """
        log.info("Refreshing Datasources %s on %s " % (str(datasource_names), self.server.server_address))

        datasource_names = [item.lower() for item in datasource_names]
        datasource_list_immutable_copy = datasource_names.copy()
        datasource_list_server = dict()
        extract_refresh_jobs = dict()

        # get all datasources from server
        for ds in TSC.Pager(self.server.datasources):
            datasource_list_server[ds.id] = ds

        # loop over server datasources and refresh if its name found in given DS list
        for ds in datasource_list_server.values():
            try:
                if ds.name.lower() in datasource_list_immutable_copy:
                    log.info('Starting extractRefresh Job for Datasource "%s" ' % ds.name)
                    refresh_job = self.refresh_extract(ds_item=ds, attempt=retry_attempt)
                    if refresh_job:
                        extract_refresh_jobs[ds.name + ':' + refresh_job.id] = refresh_job
                    datasource_names.remove(ds.name.lower())
            except Exception as e:
                log.warning(str(e).strip())

        if len(datasource_names) > 0:
            log.error("Following Datasources are not found on the server! %s " % str(datasource_names))

        if synchronous is True and len(extract_refresh_jobs) > 0:
            failed_extract_refresh_jobs = dict()
            extract_refresh_jobs_immutable_copy = extract_refresh_jobs.copy()
            log.info("Waiting for extractRefresh Jobs to Finish!")
            while len(extract_refresh_jobs) > 0:
                time.sleep(300)
                for key, _job in extract_refresh_jobs_immutable_copy.items():
                    if key in extract_refresh_jobs.keys():  # if its not yet finished!
                        time.sleep(5)
                        job = self.server.jobs.get_by_id(_job.id)
                        if job.completed_at is not None:
                            if job.finish_code == '0':
                                log.info("%s extractRefresh Succeeded %s " % (key, str(job)))
                                log.info(
                                    "%s extractRefresh Succeeded in %s " % (key, (job.completed_at - job.started_at)))
                            else:
                                log.error("%s extractRefresh Failed %s " % (key, str(job)))
                                log.error(
                                    "%s extractRefresh Failed in %s " % (key, (job.completed_at - job.started_at)))
                                failed_extract_refresh_jobs[key] = job
                            extract_refresh_jobs.pop(key)
                        else:
                            log.info("%s extractRefresh Running %s " % (key, str(job)))

            if len(failed_extract_refresh_jobs) > 0:
                raise Exception(
                    "Following extractRefresh Jobs are Failed \n[%s]!" % ','.join(failed_extract_refresh_jobs.keys()))

    def refresh_extract(self, ds_item, attempt=1, current_attempt=1):
        """

        :param ds_item:
        :param attempt:
        :param current_attempt:
        :return:
        """
        try:
            results = self.server.datasources.refresh(ds_item)
            return results
        except Exception as e:
            if current_attempt >= attempt:
                raise e
            else:
                current_attempt = current_attempt + 1
                time.sleep(5)
                log.debug(str(e))
                log.info("Refreshing '%s' failed Trying %s th time" % (ds_item.name, str(current_attempt)))
                return self.refresh_extract(ds_item=ds_item, attempt=attempt, current_attempt=current_attempt)

    def get_workbook_views(self, workbook_id):  # -> Iterable of views
        """

        :param workbook_id:
        :return:
        """
        workbook = self.server.workbooks.get_by_id(workbook_id)
        self.server.workbooks.populate_views(workbook)
        return workbook.views

    def _download_view_pdf(self, view: ViewItem, dest_dir,
                           view_filters: PDFRequestOptions = None):  # -> Filename to downloaded pdf
        log.debug("Exporting View:%s  Id:%s" % (view.name, view.id))
        Path(dest_dir).mkdir(parents=True, exist_ok=True)
        destination_filename = "%s.pdf" % os.path.join(dest_dir, view.id)
        self.server.views.populate_pdf(view_item=view, req_options=view_filters)
        with open(destination_filename, 'wb') as f:
            f.write(view.pdf)

        return destination_filename

    def download_workbook_pdf(self, workbook: WorkbookItem, dest_dir, data_filters: dict = None, page_type=None,
                              orientation=None):
        """

        :param workbook:
        :param dest_dir:
        :return:
        """
        self.server.workbooks.populate_views(workbook)

        _pdf_merger = PyPDF3.PdfFileMerger()
        _is_pdf_content_generated = False
        _pdf_file = os.path.join(dest_dir, workbook.name) + ".pdf"
        _vw_filters = PDFRequestOptions(page_type=page_type, orientation=orientation)

        if data_filters is None:
            data_filters = dict()

        for name, value in data_filters.items():
            _vw_filters.vf(name=quote_plus(name), value=quote_plus(value))

        log.info(
            "Exporting\nWorbook='%s' \nProject='%s' \nPage Type='%s' \nOrientation='%s' \nFilters='%s'\nFile='%s' " % (
                workbook.name, workbook.project_name, page_type, orientation, _vw_filters.view_filters, _pdf_file))

        for _view in workbook.views:
            _downloaded_wv = self._download_view_pdf(_view, dest_dir=os.path.join(dest_dir, 'views'),
                                                     view_filters=_vw_filters)
            _pdf_merger.append(_downloaded_wv)
            _is_pdf_content_generated = True
        if _is_pdf_content_generated:
            _pdf_merger.write(_pdf_file)
            _pdf_merger.close()
            log.info("Exported Workbook to pdf %s" % _pdf_file)
        else:
            raise Exception("No Pdf Content Generated")
        return _pdf_file

    def _download_view_png(self, view: ViewItem, dest_dir,
                           view_filters: ImageRequestOptions = None):  # -> Filename to downloaded pdf
        log.debug("Exporting View:%s  Id:%s" % (view.name, view.id))
        Path(dest_dir).mkdir(parents=True, exist_ok=True)
        destination_filename = "%s.png" % os.path.join(dest_dir, view.id)
        self.server.views.populate_image(view_item=view, req_options=view_filters)
        with open(destination_filename, 'wb') as image_file:
            image_file.write(view.image)

        return destination_filename

    def _img_concat_v_multi_resize(self, im_list, resample=Image.BICUBIC):
        min_width = min(im.width for im in im_list)
        im_list_resize = [im.resize((min_width, int(im.height * min_width / im.width)), resample=resample)
                          for im in im_list]
        total_height = sum(im.height for im in im_list_resize)
        dst = Image.new('RGB', (min_width, total_height))
        pos_y = 0
        for im in im_list_resize:
            dst.paste(im, (0, pos_y))
            pos_y += im.height
        return dst

    def download_workbook_png(self, workbook: WorkbookItem, dest_dir, data_filters: dict = None,
                              imageresolution=None,
                              maxage=-1) -> str:
        """

        :param workbook:
        :param dest_dir:
        :param data_filters:
        :param imageresolution:
        :param maxage:
        :return:
        """
        self.server.workbooks.populate_views(workbook)

        _img_list = list()
        _img_file = os.path.join(dest_dir, workbook.name) + ".png"
        _vw_filters = ImageRequestOptions(imageresolution=imageresolution, maxage=maxage)

        if data_filters is None:
            data_filters = dict()

        for name, value in data_filters.items():
            _vw_filters.vf(name=quote_plus(name), value=quote_plus(value))

        log.info(
            "Exporting\nWorbook='%s' \nProject='%s' \nFilters='%s'\nFile='%s' " % (
                workbook.name, workbook.project_name, _vw_filters.view_filters, _img_file))

        for _view in workbook.views:
            _downloaded_wv = self._download_view_png(_view, dest_dir=os.path.join(dest_dir, 'views'),
                                                     view_filters=_vw_filters)
            _img_list.append(Image.open(_downloaded_wv))

        if _img_list:
            self._img_concat_v_multi_resize(im_list=_img_list).save(_img_file)
            log.info("Exported Workbook to png %s" % _img_file)
        else:
            raise Exception("No Image Content Generated")

        return _img_file

    def _download_view_csv(self, view: ViewItem, dest_dir,
                           view_filters: CSVRequestOptions = None):  # -> Filename to downloaded pdf
        log.debug("Exporting View:%s  Id:%s" % (view.name, view.id))
        Path(dest_dir).mkdir(parents=True, exist_ok=True)
        destination_filename = "%s.csv" % os.path.join(dest_dir, view.name)
        self.server.views.populate_csv(view_item=view, req_options=view_filters)

        with open(destination_filename, 'wb') as csv_file:
            csv_file.writelines(view.csv)

        return destination_filename

    def download_workbook_csv(self, workbook: WorkbookItem, dest_dir, data_filters: dict = None) -> str:
        """

        :param workbook:
        :param dest_dir:
        :param data_filters:
        :return:
        """
        self.server.workbooks.populate_views(workbook)

        _csv_list = list()
        _excel_file = os.path.join(dest_dir, workbook.name) + ".xlsx"
        _vw_filters = CSVRequestOptions()

        if data_filters is None:
            data_filters = dict()

        for name, value in data_filters.items():
            _vw_filters.vf(name=quote_plus(name), value=quote_plus(value))

        log.info(
            "Exporting\nWorbook='%s' \nProject='%s' \nFilters='%s'\nFile='%s' " % (
                workbook.name, workbook.project_name, _vw_filters.view_filters, _excel_file))

        for _view in workbook.views:
            _downloaded_wv = self._download_view_csv(_view, dest_dir=os.path.join(dest_dir, 'views'),
                                                     view_filters=_vw_filters)
            _csv_list.append(_downloaded_wv)

        if _csv_list:
            wb = Workbook()
            # remove default work sheet
            wb.remove(wb.active)
            for _csv in _csv_list:
                _ws_name = Path(_csv).stem
                _ws = wb.create_sheet(_ws_name)
                with open(_csv) as csv_file:
                    csv_reader = csv.reader(csv_file, delimiter=',')
                    for row in csv_reader:
                        _ws.append(row)

            wb.save(_excel_file)
            wb.close()
            log.info("Exported Workbook to Excel %s" % _excel_file)
        else:
            raise Exception("No CSV Content Generated")

        return _excel_file

    def download_workbook(self, file_type: str, workbook: WorkbookItem, dest_dir, data_filters: dict = None,
                          page_type=None, orientation=None):
        if file_type.lower() == "pdf":
            return self.download_workbook_pdf(workbook=workbook, dest_dir=dest_dir, data_filters=data_filters,
                                              page_type=page_type, orientation=orientation)
        elif file_type.lower() == "png":
            return self.download_workbook_png(workbook=workbook, dest_dir=dest_dir, data_filters=data_filters)
        elif file_type.lower() == "csv":
            return self.download_workbook_csv(workbook=workbook, dest_dir=dest_dir, data_filters=data_filters)
        else:
            raise Exception("Unexpected download file_type '%s'!" % file_type)

    def _get_request_option(self, name=None, project_name=None, tag=None) -> TSC.RequestOptions:
        req_option = TSC.RequestOptions()
        if name:
            req_option.filter.add(TSC.Filter(TSC.RequestOptions.Field.Name,
                                             TSC.RequestOptions.Operator.Equals,
                                             name))
        if project_name:
            req_option.filter.add(TSC.Filter(TSC.RequestOptions.Field.ProjectName,
                                             TSC.RequestOptions.Operator.Equals,
                                             project_name))
        if tag:
            req_option.filter.add(TSC.Filter(TSC.RequestOptions.Field.Tags,
                                             TSC.RequestOptions.Operator.Equals,
                                             tag))
        return req_option

    def get_workbook_by_name(self, name, project_name=None, tag=None) -> WorkbookItem:
        """

        :param name:
        :param project_name:
        :param tag:
        :return:
        """
        req_option = self._get_request_option(name=name, project_name=project_name, tag=tag)
        all_items, pagination_item = self.server.workbooks.get(req_options=req_option)
        if not all_items:
            raise LookupError("No Workbook with given parameters (name:'%s', project:'%s', tag:'%s') found!" % (
                name, project_name, tag))
        if len(all_items) > 1:
            raise LookupError("Found multiple Workbooks with given parameters! (name:'%s', project:'%s', tag:'%s')" % (
                name, project_name, tag))

        return all_items.pop()

    def get_workbooks_by_tag(self, tag, project_name=None) -> [WorkbookItem]:
        """

        :param tag:
        :param project_name:
        :return:
        """
        req_option = self._get_request_option(tag=tag, project_name=project_name)
        all_items, pagination_item = self.server.workbooks.get(req_options=req_option)
        if not all_items:
            raise LookupError("No Workbook with given parameters found!")

        return all_items

    def get_datasource_by_name(self, name, project_name=None, tag=None) -> DatasourceItem:
        """

        :param name:
        :param project_name:
        :param tag:
        :return:
        """
        req_option = self._get_request_option(name=name, project_name=project_name, tag=tag)

        all_items, pagination_item = self.server.datasources.get(req_options=req_option)
        if not all_items:
            raise LookupError("No Datasource with given parameters found!")
        if len(all_items) > 1:
            raise LookupError("Found multiple Datasources with given parameters.")

        return all_items.pop()

    def get_project_by_name(self, name, parant_project_name=None, tag=None) -> ProjectItem:
        """

        :param name:
        :param parant_project_name:
        :param tag:
        :return:
        """
        req_option = self._get_request_option(name=name, project_name=parant_project_name, tag=tag)

        all_items, pagination_item = self.server.datasources.get(req_options=req_option)
        if not all_items:
            raise LookupError("No Project with given parameters found!")
        if len(all_items) > 1:
            raise LookupError("Found multiple Project with given parameters.")

        return all_items.pop()

    def update_all_datasource_connections(self, curr_server_address, curr_username, new_server_address=None,
                                          new_server_port=None,
                                          new_username=None, new_password=None, new_embed_password: bool = None):
        log.info('Updating "%s" Connections! server :' % str(self.server.server_address))
        for resource in TSC.Pager(self.server.datasources):
            self.server.datasources.populate_connections(resource)
            log.debug("Updating Connections for '%s' " % str(resource.name))
            for conn in resource.connections:
                if conn.server_address == curr_server_address and conn.username == curr_username:
                    log.info("Updating Connection " + str(conn))
                    if new_server_address is not None:
                        conn.server_address = new_server_address
                    if new_server_port is not None:
                        conn.server_port = new_server_port
                    if new_username is not None:
                        conn.username = new_username
                    if new_password is not None:
                        conn.password = new_password
                    if new_embed_password is not None:
                        conn.embed_password = True
                    self.server.datasources.update_connection(resource, conn)
                    log.info('Updated Connection of "%s"  server_address:%s username:%s ' % (
                        resource.name, curr_server_address, curr_username))

    def update_all_workbook_connections(self, curr_server_address, curr_username, new_server_address=None,
                                        new_server_port=None,
                                        new_username=None, new_password=None, new_embed_password: bool = None):
        log.info('Updating "%s" Connections! server :' % str(self.server.server_address))
        for resource in TSC.Pager(self.server.workbooks):
            self.server.workbooks.populate_connections(resource)
            log.debug("Updating Connections for '%s' " % str(resource.name))
            for conn in resource.connections:
                if conn.server_address == curr_server_address and conn.username == curr_username:
                    log.info("Updating Connection " + str(conn))
                    if new_server_address is not None:
                        conn.server_address = new_server_address
                    if new_server_port is not None:
                        conn.server_port = new_server_port
                    if new_username is not None:
                        conn.username = new_username
                    if new_password is not None:
                        conn.password = new_password
                    if new_embed_password is not None:
                        conn.embed_password = True
                    self.server.workbooks.update_connection(resource, conn)
                    log.info('Updated Connection of "%s"  server_address:%s username:%s ' % (
                        resource.name, curr_server_address, curr_username))


class PyTableauContent(PyTableau):
    def __init__(self, content_type, server_address, username, password, site_id, use_server_version=True,
                 verify_ssl=True):
        super().__init__(server_address, username, password, site_id, use_server_version, verify_ssl)

        endpoints = {
            'workbook': self.server.workbooks,
            'datasource': self.server.datasources,
            'subscription': self.server.subscriptions
        }
        endpoint_items = {
            'workbook': TSC.WorkbookItem,
            'datasource': TSC.DatasourceItem,
            'subscription': TSC.SubscriptionItem
        }
        endpoint_file_extensions = {
            'workbook': [".twb", ".twbx"],
            'datasource': ['.tds', '.tdsx'],
            'subscription': ['.subs']
        }
        if content_type not in endpoints:
            raise ("Accepted resource_type values are %s " % str(endpoints))
        self.endpoint = endpoints.get(content_type)
        self.endpoint_item = endpoint_items.get(content_type)
        self.endpoint_file_extensions = endpoint_file_extensions.get(content_type)


class PyTableauWorkbook(PyTableauContent):

    def __init__(self, server_address, username, password, site_id, use_server_version=True, verify_ssl=True):
        super().__init__(content_type='workbook', server_address=server_address, username=username, password=password,
                         site_id=site_id, use_server_version=use_server_version,
                         verify_ssl=verify_ssl
                         )
        self.endpoint = self.server.workbooks
        self.endpoint_item = TSC.WorkbookItem
        self.endpoint_file_extensions = [".twb", ".twbx"]


class PyTableauDatasource(PyTableauContent):

    def __init__(self, server_address, username, password, site_id, use_server_version=True, verify_ssl=True):
        super().__init__(content_type='datasource', server_address=server_address, username=username, password=password,
                         site_id=site_id, use_server_version=use_server_version,
                         verify_ssl=verify_ssl
                         )
        self.endpoint = self.server.datasources
        self.endpoint_item = TSC.DatasourceItem
        self.endpoint_file_extensions = ['.tds', '.tdsx']


class PyTableauReportScheduler():
    """

    """

    def __init__(self, tableau: PyTableau, smtp_server: smtplib.SMTP_SSL, schedule_tag, dailySchedulePrefix="Daily",
                 weeklySchedulePrefix="Weekly", monthlySchedulePrefix="Monthly"):
        self.tableau = tableau
        self.schedule_tag = schedule_tag
        self.dailySchedules = "%s" % dailySchedulePrefix
        self.weeklySchedules = "%s%s" % (weeklySchedulePrefix, str(datetime.now().isoweekday()))
        self.monthlySchedules = "%s%s" % (monthlySchedulePrefix, str(datetime.now().day))
        self.smtp_server: smtplib.SMTP_SSL = smtp_server

        try:
            log.debug(self.smtp_server.ehlo())
            log.info(self.smtp_server.helo())
        except Exception as e:
            log.error(e)
            raise e

    def __del__(self):
        try:
            self.smtp_server.quit()
        except:
            pass

    def get_scheduled_workbooks(self) -> [WorkbookItem]:
        return self.tableau.get_workbooks_by_tag(tag=self.schedule_tag)

    def _get_email_params(self, wb: WorkbookItem, schedule):
        """

        :param wb:
        :param schedule:
        :return:
        """
        email_to = list()
        email_cc = list()
        email_subject = wb.name
        _to_prefix = schedule + ':to:'
        _cc_prefix = schedule + ':cc:'
        tag: str
        for tag in wb.tags:
            if tag.startswith(_to_prefix) and '@' in tag:
                email_to.append(tag.replace(_to_prefix, ''))
            if tag.startswith(_cc_prefix) and '@' in tag:
                email_cc.append(tag.replace(_cc_prefix, ''))

        return email_subject, email_to, email_cc

    def _send_reports(self, send_from, schedule=None, email_subject=None, email_message=None,
                      data_filters: dict = None):

        log.info('Sending Reports With tag: %s:to:user@email.com ' % schedule)

        for wb in self.get_scheduled_workbooks():
            subj, to, cc = self._get_email_params(wb, schedule)
            if email_subject:
                subj = email_subject

            message = "Attached Report %s" % wb.name
            if email_message:
                message = email_message

            if to:
                log.info("Sending Workbook '%s' to: %s cc: %s" % (wb.name, COMMASPACE.join(to), COMMASPACE.join(cc)))
                self._email(wb, send_from=send_from, subj=subj, message=message, to=to, cc=cc,
                            data_filters=data_filters, file_type='pdf')

    def send_scheduled_reports(self, send_from, email_subject=None, email_message=None, data_filters: dict = None):
        """

        :param send_from:
        :param email_subject:
        :param email_message:
        """
        # self.smtp_server.connect()
        self._send_reports(send_from=send_from, schedule=self.dailySchedules, email_subject=email_subject,
                           email_message=email_message, data_filters=data_filters)
        self._send_reports(send_from=send_from, schedule=self.weeklySchedules, email_subject=email_subject,
                           email_message=email_message, data_filters=data_filters)
        self._send_reports(send_from=send_from, schedule=self.monthlySchedules, email_subject=email_subject,
                           email_message=email_message, data_filters=data_filters)

    def send_schedule(self, send_from, schedule: str, email_subject=None, email_message=None,
                      data_filters: dict = None):
        """

        :param send_from:
        :param schedule:
        :param email_subject:
        :param email_message:
        """
        # self.smtp_server.connect()
        schedule = "%s:" % schedule.strip(':')
        self._send_reports(send_from=send_from, schedule=schedule, email_subject=email_subject,
                           email_message=email_message, data_filters=data_filters)

    def send_workbook(self, wb_name, send_from: str, to: list, cc: list = None, subj: str = None, message: str = None,
                      wb_project_name=None, wb_tag=None, data_filters: dict = None, page_type=None, orientation=None,
                      file_type='pdf'):
        """

        :param wb_name:
        :param send_from:
        :param subj:
        :param message:
        :param to:
        :param cc:
        :param data_filters:
        :return:
        """
        wb = self.tableau.get_workbook_by_name(name=wb_name, project_name=wb_project_name, tag=wb_tag)
        return self._email(wb=wb, file_type=file_type, send_from=send_from, subj=subj, message=message, to=to, cc=cc,
                           data_filters=data_filters, page_type=page_type, orientation=orientation)

    def _email(self, wb, file_type, send_from: str, to: list, cc: list = None, subj: str = None, message: str = None,
               data_filters: dict = None, page_type=None, orientation=None):
        """

        :param data_filters:
        :param wb:
        :param send_from:
        :param subj:
        :param message:
        :param to:
        :param cc:
        """
        with tempfile.TemporaryDirectory() as tmpdirname:
            assert isinstance(to, list), "to is not list!"
            assert isinstance(cc, list), "cc is not list!"
            if not subj:
                subj = wb.name
            if not message:
                message = "Attached you can find %s " % wb.name

            msg = MIMEMultipart()
            msg['From'] = send_from
            msg['To'] = COMMASPACE.join(to)
            _m_to = to
            if cc:
                msg['Cc'] = COMMASPACE.join(cc)
                _m_to = _m_to + cc
            msg['Date'] = formatdate(localtime=True)
            msg['Subject'] = subj

            msg.attach(MIMEText(message))

            wb_file = self.tableau.download_workbook(file_type=file_type,
                                                     workbook=wb,
                                                     dest_dir=tmpdirname,
                                                     data_filters=data_filters,
                                                     page_type=page_type,
                                                     orientation=orientation
                                                     )
            with open(wb_file, "rb") as myfile:
                part = MIMEApplication(
                    myfile.read(),
                    Name=basename(wb_file)
                )
            # After the file is closed
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(wb_file)

            msg.attach(part)
            self.smtp_server.send_message(from_addr=self.smtp_server.user, to_addrs=_m_to, msg=msg)
            log.info("Sent Email subj:'%s' to: %s cc: %s" % (subj, COMMASPACE.join(_m_to), COMMASPACE.join(cc)))
